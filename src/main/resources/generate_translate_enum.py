import openpyxl
import re
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

I18N_PATHS = [
    "src/main/resources/i18n",
]

OUTPUT_PATH_TRANSLATE = "src/main/java/com/wis/i18n/TranslateCommon.java"
OUTPUT_PATH_KEYTRANSLATE = "src/main/java/com/wis/i18n/KeyTranslateCommon.java"
PACKAGE_NAME = "com.wis.i18n"


def to_enum_name(key: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]", "_", key)
    s = re.sub(r"_+", "_", s).upper().strip("_")
    if not s:
        s = "KEY"
    if s[0].isdigit():
        s = "K_" + s
    return s


def generate_enum_file(output_path: str, entries: dict, enum_name: str):
    if not entries:
        print(f"No entries for {enum_name}, skipped.")
        return

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"package {PACKAGE_NAME};\n\n")
        f.write("import lombok.Getter;\nimport lombok.AllArgsConstructor;\nimport lombok.ToString;\n\n")
        f.write("@Getter\n@AllArgsConstructor\n@ToString\n")
        f.write(f"public enum {enum_name} {{\n\n")

        items = sorted(entries.items())
        for i, (key, val) in enumerate(items):
            enum_name_key = to_enum_name(key)
            val = val.replace('"', '\\"')
            sep = ";" if i == len(items) - 1 else ","
            f.write(f'    {enum_name_key}("{val}"){sep}\n')

        f.write("\n    private final String description;\n}\n")

    print(f"Created file: {output_path} ({len(entries)} keys)")


def main():
    entries_translate = {}
    entries_keytranslate = {}
    total_files = 0

    for i18n_dir in I18N_PATHS:
        if not os.path.isdir(i18n_dir):
            print(f"Directory not found: {i18n_dir}")
            continue

        print(f"Scanning directory: {i18n_dir}")
        for filename in os.listdir(i18n_dir):
            if not filename.endswith(".xlsx") or filename.startswith("~$"):
                continue

            total_files += 1
            full_path = os.path.join(i18n_dir, filename)
            print(f"Reading file: {full_path}")

            try:
                wb = openpyxl.load_workbook(full_path)
                sheet = wb.active

                headers = [str(c.value).strip().lower() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
                key_idx = headers.index("key") if "key" in headers else None
                vi_idx = headers.index("vi_vn") if "vi_vn" in headers else None
                err_idx = headers.index("is_error") if "is_error" in headers else None

                if key_idx is None or vi_idx is None:
                    print(f"Missing column 'key' or 'vi_VN' in {filename}")
                    continue

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    key = row[key_idx]
                    vi_val = row[vi_idx] if vi_idx < len(row) else None
                    is_error = None
                    if err_idx is not None and err_idx < len(row):
                        is_error = str(row[err_idx]).strip().lower() if row[err_idx] is not None else None

                    if not key or not str(key).strip():
                        continue

                    key_str = str(key).strip()
                    val_str = str(vi_val).strip() if vi_val else key_str

                    # Logic: is_error == TRUE → Translate, otherwise → KeyTranslate
                    if is_error == "true":
                        entries_translate[key_str] = val_str
                    else:
                        entries_keytranslate[key_str] = val_str

            except Exception as e:
                print(f"Error reading {filename}: {e}")

    generate_enum_file(OUTPUT_PATH_TRANSLATE, entries_translate, "TranslateCommon")
    generate_enum_file(OUTPUT_PATH_KEYTRANSLATE, entries_keytranslate, "KeyTranslateCommon")

    print(f"\nSummary: {len(entries_translate)} in Translate, {len(entries_keytranslate)} in KeyTranslate from {total_files} Excel file(s).")


if __name__ == "__main__":
    main()
